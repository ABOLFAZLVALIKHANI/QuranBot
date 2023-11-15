import errno
import os
import shutil
import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand , CommandError

from django_tgbot_vip.management import helpers
from core.settings import BASE_DIR , DEBUG
from Account.models import * 

class Command(BaseCommand):
    help = 'input Words to model'

    def handle(self, *args, **options):
        try:
            
            if DEBUG:
                path =  f"{BASE_DIR}/data.xlsx"
            else:    
                path =  f"{BASE_DIR}/data.xlsx"

            excel = openpyxl.load_workbook(path)

            sheet = excel.active

            row = sheet.max_row 
            
            for i in range(2 , row ):
                number = i - 1 
                English = sheet.cell(row= i , column= 2).value
                persian = sheet.cell(row= i , column= 3).value
                ExampleEN = sheet.cell(row= i , column= 6).value
                ExampleFA = sheet.cell(row= i , column= 7).value
                # pow 
                # Example = sheet.cell(row= i , column= 7).value
                # 
                if not WordsModel.objects.filter(English = English ).exists() :

                    WordsModel.objects.create(
                        Number = number ,
                        English = English ,
                        Persian = persian ,
                        ExampleEN = ExampleEN  ,
                        ExampleFA = ExampleFA  ,
                    )

                    # DSFDFD 

        except Exception as e :
            raise CommandError(f'error for check portfo = {e}')
        
