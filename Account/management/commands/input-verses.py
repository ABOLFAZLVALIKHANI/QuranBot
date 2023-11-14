import errno
import os
import shutil
import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand , CommandError

from django_tgbot_vip.management import helpers
from core.settings import BASE_DIR , DEBUG
from Account.models import * 

class Command(BaseCommand):
    help = 'input Verses to model'

    def handle(self, *args, **options):
        try:
            
            if DEBUG:
                path =  f"{BASE_DIR}/data.xlsx"
            else:    
                path =  f"{BASE_DIR}/data.xlsx"

            excel = openpyxl.load_workbook(path)

            sheet = excel.active

            row = sheet.max_row 
            
            for i in range(3 , row ):
                number = i - 1 
                arabic = sheet.cell(row= i , column= 2).value
                persian = sheet.cell(row= i , column= 3).value
                jose = sheet.cell(row= i , column= 5).value
                soreh = sheet.cell(row= i , column= 7).value
                
                if not VersesModel.objects.filter(Arabic = arabic ).exists() :
                    VersesModel.objects.create(
                        Number = number ,
                        Persian = persian ,
                        Arabic = arabic ,
                        Soreh = soreh ,
                        Jose = jose 
                    )
        except Exception as e :
            raise CommandError(f'error for check portfo = {e}')
        
